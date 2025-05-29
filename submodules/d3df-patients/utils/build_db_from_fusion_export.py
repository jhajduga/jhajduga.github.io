import pandas as pd
from pathlib import Path
import os
import argparse
from loguru import logger as logging
import sys
import datetime
import time
import hashlib
import re
import json

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


def compute_hash(row):
    """
    Computes a SHA256 hash based on:
      - SC id
      - Slice id
      - Fiber id
      - SC box id
      - Experiment Date

    If any field is missing or None, an empty string is used.
    """
    sc_id = str(row.get('SC id', '') or '')
    slice_id = str(row.get('Slice id', '') or '')
    fiber_id = str(row.get('Fiber id', '') or '')
    sc_box_id = str(row.get('SC box id', '') or '')
    exp_date = str(row.get('Experiment Date', '') or '')
    s = f"{sc_id}_{slice_id}_{fiber_id}_{sc_box_id}_{exp_date}"
    return hashlib.sha256(s.encode('utf-8')).hexdigest()


def normalize_column_name(name: str) -> str:
    """
    Normalizes a column name to a uniform convention:
      - Strips whitespace and removes any leading '#' or BOM.
      - Inserts an underscore between a lowercase letter and an uppercase letter
        (e.g. "ComponentName" becomes "Component_Name").
      - Replaces any whitespace or hyphen with an underscore.
      - Replaces multiple underscores with a single underscore.
      - Capitalizes each word.
    
    Examples:
      "ComponentName"  -> "Component_Name"
      " Body Name "    -> "Body_Name"
    """
    name = name.strip().lstrip('#').lstrip('\ufeff')
    # Insert underscore between a lowercase letter and an uppercase letter
    name = re.sub(r'(?<=[a-z])([A-Z])', r'_\1', name)
    # Replace spaces and hyphens with underscore
    name = re.sub(r'[\s\-]+', '_', name)
    # Replace multiple underscores with a single underscore
    name = re.sub(r'_+', '_', name)
    # Capitalize each word
    parts = name.split('_')
    parts = [p.capitalize() for p in parts]
    return '_'.join(parts)


class D3DFDataDB:
    def __init__(self, input_path: os.PathLike, output_filename='d3df_scintillator_mapping_db.xlsx'):
        self.input_path = Path(input_path).resolve()
        logging.info(f"Input path: {self.input_path}")
        self.output_filename = output_filename
        self.df = pd.DataFrame()

    def read_df_from_csv(self, sep=';'):
        """
        Reads a CSV file where the first row is treated as the header.
        The header values are normalized using normalize_column_name.
        """
        # Read CSV without header
        self.df = pd.read_csv(self.input_path, sep=sep, header=None)
        # Use the first row as header and normalize names
        new_header = self.df.iloc[0].apply(lambda x: normalize_column_name(str(x)))
        self.df = self.df[1:].copy()
        self.df.columns = new_header

        # Strip whitespace from all string columns
        for col_name in self.df.columns:
            try:
                self.df[col_name] = self.df[col_name].astype(str).str.strip()
            except Exception as e:
                logging.warning(f"Error stripping column {col_name}: {e}")
        logging.info(f"CSV loaded with {len(self.df)} rows and columns: {self.df.columns.tolist()}")
        return self.df

    def add_extra_columns(self):
        """
        Adds extra columns (to be filled manually later) if they do not exist.
        Missing numeric values are set to pd.NA.
        """
        extra_cols = {
            'Slice_id': pd.NA,
            'PMT_chn': pd.NA,
            'SC_id': pd.NA,
            'SC_wrap_id': pd.NA,
            'Fiber_id': pd.NA,
            'PP_id': pd.NA,
            'Cal_params': "",
            'Cal_fun': "",
            'Date': pd.NA
        }
        for col, default in extra_cols.items():
            if col not in self.df.columns:
                self.df[col] = default

    def sort_by_component(self):
        """
        Sorts the DataFrame by the component name.
        Checks for either "Component_Name" or "Componentname".
        """
        if "Component_Name" in self.df.columns:
            self.df.sort_values(by="Component_Name", inplace=True)
            self.df.reset_index(drop=True, inplace=True)
        elif "Componentname" in self.df.columns:
            self.df.sort_values(by="Componentname", inplace=True)
            self.df.reset_index(drop=True, inplace=True)
        else:
            logging.warning("Column 'Component_Name' (or 'Componentname') does not exist – cannot sort by component.")

    def add_hash_column(self):
        """
        Adds a new column 'Hash' by applying compute_hash on each row.
        """
        self.df['Hash'] = self.df.apply(compute_hash, axis=1)

    def reorder_columns(self):
        """
        Reorders the columns according to a predefined order.
        The desired order is stored in a single list; extra columns are ignored.
        """
        # Define the desired column order exactly matching the normalized names from the CSV
        desired_order = [
            "Hash",
            "Component_Name", "Body_Name",
            "Co_M_X", "Co_M_Y", "Co_M_Z",
            "SC_id", "SC_wrap_id", "Fiber_id", "PMT_chn",
            "Slice_id", "PP_id",
            "Bb_Min_X", "Bb_Min_Y", "Bb_Min_Z",
            "Bb_Max_X", "Bb_Max_Y", "Bb_Max_Z",
            "Appearance", "Physical_Material",
            "Body_Vertices", "Mesh_Nodes", "Mesh_Vertices", "Mesh_Normal_Vectors",
            "Cal_params", "Cal_fun", "Date"
        ]
        # Use only those columns that exist in the DataFrame:
        final_order = [col for col in desired_order if col in self.df.columns]
        self.df = self.df.reindex(columns=final_order)
        logging.info(f"Final column order: {self.df.columns.tolist()}")

    def save_to_excel(self, sheet_data='scintillator_mapping_db', sheet_meta='Metadata'):
        """
        Saves the DataFrame to an Excel file with two sheets:
        - Data sheet (main data formatted as an Excel table).
        - Metadata sheet (key-value pairs with additional experiment metadata).

        Missing values in data are replaced with empty strings.
        Certain columns specified in a list are hidden.
        """
        self.reorder_columns()
        # Replace missing values (pd.NA) with empty strings to ensure compatibility with openpyxl
        self.df = self.df.fillna("")

        # Create workbook and data sheet
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = sheet_data

        # Write DataFrame rows to Excel sheet
        for row in dataframe_to_rows(self.df, index=False, header=True):
            ws_data.append(row)

        # Define Excel table area
        max_row, max_col = ws_data.max_row, ws_data.max_column
        table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table_style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        table = Table(displayName=f"D3DF_Scintillator_{int(time.time())}", ref=table_ref)
        table.tableStyleInfo = table_style
        ws_data.add_table(table)

        # Adjust column widths based on the maximum content length
        for col in ws_data.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            adjusted_width = max_length + 6  # extra spacing
            ws_data.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        # Hide specified columns
        columns_to_hide = ['Appearance', 'Cal_fun', 'Bb_Min_X', 'Bb_Min_Y', 'Bb_Min_Z', 
                        'Bb_Max_X', 'Bb_Max_Y', 'Bb_Max_Z', 'Physical_Material', 
                        'Body_Vertices', 'Mesh_Nodes', 'Mesh_Vertices', 'Mesh_Normal_Vectors']

        for idx, col in enumerate(self.df.columns, start=1):
            if col in columns_to_hide:
                ws_data.column_dimensions[get_column_letter(idx)].hidden = True

        # Create metadata sheet
        ws_meta = wb.create_sheet(title=sheet_meta)
        
        # Base metadata (automatically generated)
        metadata = {
            'Data CSV': str(self.input_path),
            'Generated Date': datetime.datetime.now().isoformat(),
            'Experiment Date': '',  # To be filled manually
            'Treatment Plan ID': '',  # To be filled manually
            'Planned Dose (Gy)': '',  # To be filled manually
            'Patient Position': '',  # To be filled manually
            'Table Position (mm)': '',  # To be filled manually
            'Isocenter Position (mm)': '',  # To be filled manually
            'Phantom Transformation Matrix': ''  # To be filled manually
        }

        # Write metadata to Excel as key-value pairs
        ws_meta.append(["Key", "Value"])
        for key, value in metadata.items():
            ws_meta.append([key, value])

        # Adjust metadata column widths
        for col in ws_meta.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            adjusted_width = max_length + 6
            ws_meta.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        # Save workbook
        wb.save(self.output_filename)
        logging.info(f"Saved data and metadata to Excel file: {self.output_filename}")

    def translate_material_names(self, json_path: str):
            """
            Translates material names in the 'Physical_Material' column using a provided JSON mapping.
            Unmapped values are left unchanged.
            """
            if not json_path or not Path(json_path).exists():
                logging.warning("No valid material translation JSON provided. Skipping translation.")
                return
            
            try:
                with open(json_path, "r", encoding="utf-8") as f:
                    material_map = json.load(f)

                if "Physical_Material" in self.df.columns:
                    original_unique = self.df["Physical_Material"].unique()
                    self.df["Physical_Material"] = self.df["Physical_Material"].map(material_map).fillna(self.df["Physical_Material"])
                    translated_unique = self.df["Physical_Material"].unique()
                    logging.info(f"Translated materials. Unique values before: {list(original_unique)}")
                    logging.info(f"Unique values after translation: {list(translated_unique)}")
                else:
                    logging.warning("Column 'Physical_Material' not found. No translation performed.")

            except Exception as e:
                logging.error(f"Error during material translation: {e}")


def main():

    parser = argparse.ArgumentParser(
        description="Reads a CSV file with 3D data and saves it to an Excel file (two sheets: data and metadata)."
    )
    parser.add_argument(
        "--csv_path",
        type=str,
        default="./D3DF_bodies.csv",
        help="Path to the CSV file (default: ./D3DF_bodies.csv)"
    )
    parser.add_argument(
        "--output",
        type=str,
        default="./d3df_scintillator_mapping_db.xlsx",
        help="Name of the output Excel file (default: ./d3df_scintillator_mapping_db.xlsx)"
    )
    parser.add_argument(
    "--material_map",
    type=str,
    default=None,
    help="Optional path to a JSON file with material name translation."
)

    args = parser.parse_args()
    
    
    logging.remove()
    logging.add(sys.stderr,
               format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {message}",
               level="INFO")
    logging.info("Starting conversion...")
    logging.info(f"Input path: {args.csv_path}")
    db = D3DFDataDB(input_path=args.csv_path, output_filename=args.output)
    db.read_df_from_csv(sep=';')
    db.add_extra_columns()
    db.sort_by_component()  # Sort by component if possible.
    if args.material_map:
        db.translate_material_names(args.material_map)
    db.add_hash_column()
    db.save_to_excel()


if __name__ == "__main__":
    main()
